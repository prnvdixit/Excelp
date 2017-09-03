import openpyxl as xl
import os
import sys

from openpyxl.chart import PieChart3D as pie3
from openpyxl.chart import Reference as ref

def get_column_codes(column_num):

    """
    This function takes in a integer (number) and returns the corresponding
    excel-column-code. eg. 27 - "AA".
    """

    column_code = ""

    while(column_num > 0):
        column_code = chr((column_num - 1) % 26 + ord('A')) + column_code
        column_num = (column_num - 1) / 26

    return column_code



def compare_sheets(orig, comp):
    
    """
    Utility :- Teachers who teach Finance and are expected to grade excel sheets
    can call this function repeatedly with 'orig' being the transcript file and 
    'comp' being the student-answer-sheet.

    This function returns if the two excel files - 'orig' and 'comp' are
    having same number of sheets and same values in each cell of each sheet.
    In case, even the number of sheets aren't same or size of one of the file
    has more cells filled - Respective error message is shown.
    If the values at two cells are different - It shows the same corresponding
    to 'orig' excel file.
    """

    try:
        # Load the workbook if it is already present in the file directory
        wb = xl.load_workbook(orig)
        wb2 = xl.load_workbook(comp)

    except:
        # Otherwise make the new files
        wb = xl.Workbook()
        wb.save(orig)
        wb2 = xl.Workbook()
        wb2.save(comp)

    
    # Store the names of sheets in arrays
    sheet_names_orig = wb.get_sheet_names()
    sheet_names_comp = wb2.get_sheet_names()

    # Number of sheets in each of the files
    num_sheets_orig = len(sheet_names_orig)
    num_sheets_comp = len(sheet_names_comp)

    
    # If the number of sheets are different - Show error message and exit
    if num_sheets_orig != num_sheets_comp :
        print "Spreadsheets are not even having same number of sheets"
        sys.exit()


    # Now iterate through all the sheets
    for sheet_no in xrange(num_sheets_orig):
        max_row_orig = wb[sheet_names_orig[sheet_no]].max_row
        max_row_comp = wb2[sheet_names_comp[sheet_no]].max_row

        max_column_orig = wb[sheet_names_orig[sheet_no]].max_column
        max_column_comp = wb2[sheet_names_comp[sheet_no]].max_column

        # If the Maximum row filled in any of the sheets is not same as that corresponding 
        # to other sheet - Exit after showing corresponding error message
        if max_row_orig != max_row_comp or max_column_orig != max_column_comp:
            print "Some rows/column in one of the sheets are extra filled"
            sys.exit()

        # Corresponding to each sheet, iterate through each cell
        for column in xrange(1, max_column_orig + 1):
            for row in xrange(1, max_row_orig + 1):

                # Obtain cell code corresponding to the given column number
                cell_code = get_column_codes(column) + str(row)

                # If value in any cell is different in two sheets - Print corresponding message
                if wb[sheet_names_orig[sheet_no]][cell_code].value != wb2[sheet_names_comp[sheet_no]][cell_code].value:
                    print "Sheet \"" + sheet_names_comp[sheet_no] + "\" is storing different value at", cell_code



def merge_sheets(orig):
    
    """
    Utility :- MS Excel already supports joining different excel files
    into one file. This is particularly useful in any large project when
    a lot of people are involved in fetching parts of data. But if after
    that, one needs to join all the sheets of a file into one sheet, he
    can call this function with the 'orig' being the file whose sheets
    have to be merged.

    This function modifies the 'orig' file and merge all the sheets in it
    into the first sheet.
    """

    try:
        # Load the workbook if it is already present in the file directory        
        wb = xl.load_workbook(orig)

    except:
        # Else make the new file        
        wb = xl.Workbook()
        wb.save(comp)

    # Stores all the names of sheets
    sheet_names = wb.get_sheet_names()

    # Stores name of the first sheet - In which, all others are to be merged
    main_sheet = sheet_names[0]

    for merged_sheets in sheet_names[1:]:

        # Data is a 2D representation of the data in the sheets
        data = [[cell.value for cell in rows] for rows in wb[merged_sheets].rows]

        # Append the data, row-by-row into the main sheet of the workbook
        for data_row in data:
            wb[main_sheet].append(data_row)

    # Save the modified workbook
    wb.save(orig)



def visualise_pie_column(orig):

    """
    Utility :- Visualising Data using 3D pie-charts. This can be used to
    plot pie-chart for a single column with appropriate labels and leaders.

    This function requires the user to supply the column-number (numbering 
    starting from 1, but not 1 as the labels are usually present in column -1).
    Finally, it adds the pie-chart built from the data to the horizontal end of file.
    """

    try:
        # Load the workbook if it is already present in the file directory
        wb = xl.load_workbook(orig)

    except:
        # Else make the new file
        wb = xl.Workbook()
        wb.save(orig)

    # After this, ws will be equivalent to wb["Sheet"] i.e. first sheet of file
    ws = wb.active

    print "Enter Column-number for visualisation"
    print "{Column-number would be greater than 1}"
    
    # Ask the user for column-number
    column_num = int(raw_input())

    # Get the excel-column-code from the respective function
    column_code = get_column_codes(column_num)

    # Find the Maximum row number
    max_row_no = ws.max_row

    # Initialize the 3D Pie-Chart object
    pie = pie3()

    # Labels are generally found in the first column, second row onwards
    labels = ref(ws, min_col = 1, min_row = 2, max_row = max_row_no)

    # Data is usually from the first row (leader of data is in first row) to last
    data = ref(ws, min_col = column_num, min_row = 1, max_row = max_row_no)

    # Create the respective Pie chart by adding data and labels to it
    pie.add_data(data, titles_from_data = True)
    pie.set_categories(labels)
    
    # Get location of the Next column after the last column in terms of excel-column-code
    next_to_max_cell = get_column_codes(ws.max_column + 1) + str(1)

    # Finally, insert the chart into the workbook
    ws.add_chart(pie, next_to_max_cell)

    # Save the modified workbook
    wb.save(orig)


if __name__ == "__main__":
    orig = "/home/prnvdixit/Desktop/orig.xlsx"
    comp = "/home/prnvdixit/Desktop/comp.xlsx"

    merge_sheets(orig)    
    compare_sheets(orig, comp)
    visualise_pie_column(orig)
